from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLOUR_MAP = {
    "MATCH": "C6EFCE",
    "QTY MISMATCH": "FFEB9C",
    "MISSING IN TARGET": "FFC7CE",
    "EXTRA IN TARGET": "FFD18C",
    "default": "FFFFFF",
}


def _build_width_map(original_target: Any, sheet_name: str | None = None) -> dict[str, float]:
    if original_target is None:
        return {}

    try:
        if hasattr(original_target, "getvalue"):
            source_stream = BytesIO(original_target.getvalue())
        elif isinstance(original_target, (bytes, bytearray)):
            source_stream = BytesIO(original_target)
        else:
            source_stream = original_target

        wb = load_workbook(source_stream)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        header_to_width: dict[str, float] = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            header = str(cell.value).strip() if cell.value is not None else ""
            if not header:
                continue
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width is not None:
                header_to_width[header] = width
        return header_to_width
    except Exception:
        return {}


def _remark_fill(remark: str) -> PatternFill:
    text = (remark or "").upper()
    if "QTY MISMATCH" in text:
        color = COLOUR_MAP["QTY MISMATCH"]
    elif "MISSING IN TARGET" in text:
        color = COLOUR_MAP["MISSING IN TARGET"]
    elif "EXTRA IN TARGET" in text:
        color = COLOUR_MAP["EXTRA IN TARGET"]
    elif "MATCH" in text:
        color = COLOUR_MAP["MATCH"]
    else:
        color = COLOUR_MAP["default"]
    return PatternFill(fill_type="solid", start_color=color, end_color=color)


def write_annotated_excel(
    df: pd.DataFrame,
    original_target_path: Any = None,
    sheet_name: str = "Compared_Output",
) -> BytesIO:
    """
    Write DataFrame to BytesIO and apply Remarks formatting.
    """
    if "Remarks" not in df.columns:
        df = df.copy()
        df["Remarks"] = ""

    width_map = _build_width_map(original_target_path, sheet_name=sheet_name)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    output.seek(0)
    wb = load_workbook(output)
    ws = wb[sheet_name]

    header_row = [cell.value for cell in ws[1]]
    remarks_col_idx = header_row.index("Remarks") + 1

    ws.cell(row=1, column=remarks_col_idx).font = Font(bold=True)

    max_len = len("Remarks")
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=remarks_col_idx)
        remark = "" if cell.value is None else str(cell.value)
        cell.fill = _remark_fill(remark)
        max_len = max(max_len, len(remark))

    remarks_letter = get_column_letter(remarks_col_idx)
    ws.column_dimensions[remarks_letter].width = min(max_len + 4, 120)

    for col_idx, header in enumerate(header_row, start=1):
        if header in width_map:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width_map[header]

    ws.freeze_panes = "A2"

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output
